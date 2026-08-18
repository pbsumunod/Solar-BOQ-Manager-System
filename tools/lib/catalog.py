"""Materials catalog: a reusable, editable, multi-store list of
materials/prices that BOQ line items can be pulled from, independent of any
single project. Each catalog row belongs to exactly one store (via "Store
Name"), so the same Material can appear multiple times -- once per store
that carries it, each with its own price.

Kept free of any Streamlit imports, same as boq_data.py, so it can be tested
and reused (e.g. by the standalone import script) on its own.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
CATALOG_PATH = PROJECT_ROOT / "data" / "materials_catalog.xlsx"
STORES_PATH = PROJECT_ROOT / "data" / "stores.xlsx"
CATEGORY_LIST_PATH = PROJECT_ROOT / "data" / "category_list.xlsx"

CATALOG_COLUMNS = ["Category", "Material", "Brand", "Model", "Unit of Measurement", "Store Name", "Unit Cost (₱)"]
STORES_COLUMNS = ["Store Name", "Address"]
CATEGORY_LIST_COLUMNS = ["Category"]

# The store name new/legacy Materials rows default to when their actual
# source store is unknown -- deliberately NOT a real store name (e.g. not
# "Plug and Go"), since that would falsely assert provenance and would
# wrongly make the row eligible for the Store->Unit Cost auto-lookup in
# boq_data.apply_store_pricing() against a store it was never priced from.
CUSTOM_STORE_LABEL = "— Custom / Manual —"

# Collapse known spelling/casing variants of the same physical unit down to
# one canonical string, so "pc"/"Pc"/"pcs" (etc) never coexist in the
# catalog as if they were different units.
UNIT_ALIASES = {
    "pc": "pcs",
    "pcs": "pcs",
    "piece": "pcs",
    "pieces": "pcs",
    "mtr": "m",
    "mtrs": "m",
    "meter": "m",
    "meters": "m",
    "metre": "m",
    "metres": "m",
}

# Rows below a "Description"/category cell containing any of these phrases
# are cost-summary/labor lines, not materials -- stop importing a sheet once
# one of these is hit.
STOP_PHRASES = (
    "total material cost",
    "mark up on materials",
    "sign and seal",
    "installation",
    "engineering",
    "transportation",
    "total service cost",
    "total project cost",
)


def _clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_unit(value) -> str:
    """Map a unit-of-measurement string to its single canonical spelling."""
    cleaned = _clean(value)
    if not cleaned:
        return ""
    return UNIT_ALIASES.get(cleaned.lower(), cleaned.lower())


# ---------------------------------------------------------------------------
# Catalog (multi-store)
# ---------------------------------------------------------------------------

def blank_catalog_df() -> pd.DataFrame:
    return pd.DataFrame(columns=CATALOG_COLUMNS)


def normalize_catalog_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in CATALOG_COLUMNS:
        if col not in df.columns:
            df[col] = "" if col != "Unit Cost (₱)" else 0.0
    df["Unit Cost (₱)"] = pd.to_numeric(df["Unit Cost (₱)"], errors="coerce").fillna(0.0).astype(float)
    for col in ("Category", "Material", "Brand", "Model", "Store Name"):
        df[col] = df[col].fillna("").astype(str)
    df["Unit of Measurement"] = df["Unit of Measurement"].apply(normalize_unit)
    return df[CATALOG_COLUMNS].reset_index(drop=True)


def load_catalog() -> pd.DataFrame:
    if not CATALOG_PATH.exists():
        return blank_catalog_df()
    df = pd.read_excel(CATALOG_PATH, sheet_name="Catalog", engine="openpyxl")
    return normalize_catalog_df(df)


def save_catalog(df: pd.DataFrame) -> Path:
    CATALOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    df = normalize_catalog_df(df)
    df.to_excel(CATALOG_PATH, sheet_name="Catalog", index=False, engine="openpyxl")
    return CATALOG_PATH


def parse_package_workbook(path: str | Path, store_name: str = "Plug and Go") -> pd.DataFrame:
    """Extract catalog rows from a hierarchical package BOQ workbook.

    Expects the layout used by files like PACKAGES.xlsx: a header row
    ("No., Description, Technical Specifications, Brand/Model, Quantity,
    UoM, Price, Unit Cost, Total Cost (PHP)"), followed by material rows
    where the Description column is only filled on the first row of a new
    category and blank on continuation rows for the same category. Stops
    at the first cost-summary row (Total Material Cost, Installation, etc).

    The source "Brand/Model" column is a mix of real brand names, model
    numbers, and "X or equivalent" notes -- this generic parser can't
    reliably tell those apart, so it's dropped into Brand as-is and Model
    is left blank. Review/split Brand vs Model by hand afterward (in the
    Materials Catalog tab) for anything imported this way.

    The workbook itself has no concept of "store" -- every row parsed from
    it is attributed to `store_name` (the store this particular quote/price
    list came from), defaulting to "Plug and Go" since that's the store the
    very first package quote this parser was built for (PACKAGES.xlsx) came
    from.
    """
    wb = load_workbook(path, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and _clean(row[0]).lower() == "no.":
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        category = ""
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            description = _clean(row[1]) if len(row) > 1 else ""
            if description:
                if any(phrase in description.lower() for phrase in STOP_PHRASES):
                    break
                category = description

            spec = _clean(row[2]) if len(row) > 2 else ""
            brand = _clean(row[3]) if len(row) > 3 else ""
            uom = normalize_unit(row[5]) if len(row) > 5 else ""
            unit_cost = row[7] if len(row) > 7 else None

            material = spec or description
            if not material or not category:
                continue
            if not isinstance(unit_cost, (int, float)):
                continue

            records.append(
                {
                    "Category": category,
                    "Material": material,
                    "Brand": brand,
                    "Model": "",
                    "Unit of Measurement": uom,
                    "Store Name": store_name,
                    "Unit Cost (₱)": float(unit_cost),
                }
            )

    if not records:
        return blank_catalog_df()

    df = pd.DataFrame(records, columns=CATALOG_COLUMNS)
    df = df.drop_duplicates(
        subset=["Category", "Material", "Brand", "Unit of Measurement", "Store Name"]
    ).reset_index(drop=True)
    return df


def merge_into_catalog(new_items_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Merge new catalog rows into the existing catalog, skipping exact
    (Category, Material, Brand, Unit of Measurement, Store Name) duplicates.
    Returns (merged_df, number_of_rows_added)."""
    existing = load_catalog()
    key_cols = ["Category", "Material", "Brand", "Unit of Measurement", "Store Name"]
    existing_keys = set(map(tuple, existing[key_cols].values.tolist()))

    rows_to_add = [
        row for row in new_items_df.to_dict("records")
        if tuple(row[c] for c in key_cols) not in existing_keys
    ]

    if not rows_to_add:
        return existing, 0

    new_df = pd.DataFrame(rows_to_add, columns=CATALOG_COLUMNS)
    merged = new_df if existing.empty else pd.concat([existing, new_df], ignore_index=True)
    return merged, len(rows_to_add)


# ---------------------------------------------------------------------------
# Stores (Store Name + Address, unique on Store Name)
# ---------------------------------------------------------------------------

def blank_stores_df() -> pd.DataFrame:
    return pd.DataFrame(columns=STORES_COLUMNS)


def normalize_stores_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in STORES_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    for col in STORES_COLUMNS:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df = df[df["Store Name"] != ""]
    # Case-insensitive dedup, same rationale as normalize_simple_list_df.
    df = df.loc[~df["Store Name"].str.lower().duplicated()]
    return df[STORES_COLUMNS].reset_index(drop=True)


def load_stores() -> pd.DataFrame:
    if not STORES_PATH.exists():
        return blank_stores_df()
    df = pd.read_excel(STORES_PATH, sheet_name="Stores", engine="openpyxl")
    return normalize_stores_df(df)


def save_stores(df: pd.DataFrame) -> Path:
    STORES_PATH.parent.mkdir(parents=True, exist_ok=True)
    df = normalize_stores_df(df)
    df.to_excel(STORES_PATH, sheet_name="Stores", index=False, engine="openpyxl")
    return STORES_PATH


def ensure_store_exists(store_name: str, address: str = "") -> None:
    """Idempotent: append store_name to Stores if not already present.
    Used by the workbook importer and the "Add new store" UI flow so a
    store name that appears on catalog rows always has a matching Stores
    entry (keeps dropdown options in sync with what's actually in use)."""
    store_name = store_name.strip()
    if not store_name:
        return
    stores_df = load_stores()
    if store_name.lower() in {s.lower() for s in stores_df["Store Name"]}:
        return
    stores_df = pd.concat(
        [stores_df, pd.DataFrame([{"Store Name": store_name, "Address": address}])], ignore_index=True
    )
    save_stores(stores_df)


# ---------------------------------------------------------------------------
# Category list / Material list -- flat, independent, editable master lists.
# Structurally identical (one text column, dedup, sort), so both share
# small internal helpers.
# ---------------------------------------------------------------------------

def normalize_simple_list_df(df: pd.DataFrame, column_name: str) -> pd.DataFrame:
    df = df.copy()
    if column_name not in df.columns:
        df[column_name] = ""
    df[column_name] = df[column_name].fillna("").astype(str).str.strip()
    df = df[df[column_name] != ""]
    # Dedup case-insensitively (keeping the first occurrence's casing) so
    # "Panels" and "panels" don't both survive as separate list entries.
    df = df.loc[~df[column_name].str.lower().duplicated()]
    df = df.sort_values(column_name, key=lambda s: s.str.lower())
    return df[[column_name]].reset_index(drop=True)


def _load_simple_list(path: Path, sheet_name: str, column_name: str) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=[column_name])
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    return normalize_simple_list_df(df, column_name)


def _save_simple_list(df: pd.DataFrame, path: Path, sheet_name: str, column_name: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    df = normalize_simple_list_df(df, column_name)
    df.to_excel(path, sheet_name=sheet_name, index=False, engine="openpyxl")
    return path


def blank_category_list_df() -> pd.DataFrame:
    return pd.DataFrame(columns=CATEGORY_LIST_COLUMNS)


def load_category_list() -> pd.DataFrame:
    return _load_simple_list(CATEGORY_LIST_PATH, "Categories", "Category")


def save_category_list(df: pd.DataFrame) -> Path:
    return _save_simple_list(df, CATEGORY_LIST_PATH, "Categories", "Category")


def sync_category_list_from_catalog(catalog_df: pd.DataFrame, category_list_df: pd.DataFrame) -> pd.DataFrame:
    """Ensure every Category present in catalog_df has a corresponding
    entry in the Category list -- adding whatever's missing, never
    removing or overwriting anything already there. Shared by the catalog
    importer (tools/import_catalog_from_workbook.py) and the multi-store
    migration script, so a newly imported category is immediately
    selectable in the Category dropdown, not just present in the flat
    catalog."""
    existing_cats = {c.lower() for c in category_list_df["Category"]}
    new_cats = sorted({c for c in catalog_df["Category"] if c and c.lower() not in existing_cats})
    if new_cats:
        category_list_df = pd.concat(
            [category_list_df, pd.DataFrame({"Category": new_cats})], ignore_index=True
        )
        category_list_df = normalize_simple_list_df(category_list_df, "Category")
    return category_list_df


def distinct_materials(catalog_df: pd.DataFrame) -> list:
    """Sorted, deduped (case-insensitively) list of Material names
    currently present in the catalog -- this *is* the material list now;
    there's no separate master list to maintain in sync with it, since
    every material that matters already has at least one catalog row."""
    if catalog_df is None or catalog_df.empty or "Material" not in catalog_df.columns:
        return []
    seen: dict = {}
    for value in catalog_df["Material"]:
        cleaned = str(value).strip()
        if cleaned and cleaned.lower() not in seen:
            seen[cleaned.lower()] = cleaned
    return sorted(seen.values(), key=str.lower)


def material_category_map(catalog_df: pd.DataFrame) -> dict:
    """Material -> Category lookup, derived directly from the catalog
    (first occurrence wins for a given Material, consistent across
    however many stores carry it). This is the Category<->Material
    "mapping" in full -- there's no separate stored representation of it."""
    if catalog_df is None or catalog_df.empty:
        return {}
    mapping: dict = {}
    for _, row in catalog_df.iterrows():
        material = str(row.get("Material", "")).strip()
        category = str(row.get("Category", "")).strip()
        if material and material not in mapping and category:
            mapping[material] = category
    return mapping


def sort_catalog_df(df: pd.DataFrame) -> pd.DataFrame:
    """Default browse order for the catalog: Category, then Material, then
    Unit Cost, all ascending. Text keys sort case-insensitively (so
    "panels" and "Panels" land together) via a throwaway lowercase sort
    key rather than mutating the real columns. Used both for the Materials
    Catalog table itself and anything built from it (the "Add items from
    catalog" picker), so browsing and picking always agree on order."""
    if df is None or df.empty:
        return df
    df = df.copy()
    df["_cat"] = df["Category"].astype(str).str.lower()
    df["_mat"] = df["Material"].astype(str).str.lower()
    df["_cost"] = pd.to_numeric(df["Unit Cost (₱)"], errors="coerce")
    df = df.sort_values(["_cat", "_mat", "_cost"], kind="mergesort")
    return df.drop(columns=["_cat", "_mat", "_cost"])


def cheapest_catalog_rows(catalog_df: pd.DataFrame) -> pd.DataFrame:
    """One catalog row per distinct Material (case-insensitive) -- whichever
    store has the lowest Unit Cost for that Material. Ties keep whichever
    row appears first. Powers the "Add all materials (lowest price)" bulk
    -add action, so adding the whole catalog to a BOQ in one click never
    adds the same Material twice at an inflated price just because a
    pricier store's row happened to be picked."""
    if catalog_df is None or catalog_df.empty:
        return catalog_df
    df = catalog_df.copy()
    df["_key"] = df["Material"].astype(str).str.strip().str.lower()
    df = df[df["_key"] != ""]
    if df.empty:
        return df.drop(columns=["_key"])
    df["_cost"] = pd.to_numeric(df["Unit Cost (₱)"], errors="coerce")
    cheapest_idx = df.groupby("_key")["_cost"].idxmin()
    return df.loc[cheapest_idx].drop(columns=["_key", "_cost"])
