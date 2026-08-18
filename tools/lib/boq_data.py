"""BOQ project data layer: schemas, persistence, template generation, validation.

Kept free of any Streamlit imports so it can be tested and reused on its own
(e.g. from a future non-UI tool that just needs to read/write a project file).
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import catalog

PROJECT_ROOT = Path(__file__).resolve().parents[2]
PROJECTS_DIR = PROJECT_ROOT / "data" / "projects"

SCHEMA_VERSION = "1"

CORE_MATERIAL_COLUMNS = [
    "Category",
    "Material",
    "Brand",
    "Quantity",
    "Unit of Measurement",
    "Store Name",
    "Unit Cost (₱)",
    "Total Cost (₱)",
    "Notes",
]

# Total Cost is always derived (Quantity x Unit Cost), never trusted as
# user-entered input -- see recompute_material_totals().
REQUIRED_MATERIAL_COLUMNS = [
    "Material",
    "Brand",
    "Quantity",
    "Unit of Measurement",
    "Unit Cost (₱)",
]

NUMERIC_MATERIAL_COLUMNS = ["Quantity", "Unit Cost (₱)", "Total Cost (₱)"]

# Columns that used to be part of the core schema and were deliberately
# removed -- drop them outright when loading older project files instead of
# resurrecting them as a "custom" column.
LEGACY_DROPPED_MATERIAL_COLUMNS = ["Supplier / Store"]

CORE_EXPENSE_COLUMNS = ["Expense Name", "Amount (₱)", "Notes"]
NUMERIC_EXPENSE_COLUMNS = ["Amount (₱)"]

# Header aliases -> canonical core column name, matched case-insensitively
# after stripping whitespace, so uploads that don't match the template
# header-for-header still work.
COLUMN_ALIASES = {
    "material": "Material",
    "item": "Material",
    "brand": "Brand",
    "brand of material": "Brand",
    "qty": "Quantity",
    "quantity": "Quantity",
    "unit": "Unit of Measurement",
    "unit of measure": "Unit of Measurement",
    "unit of measurement": "Unit of Measurement",
    "uom": "Unit of Measurement",
    "unit cost": "Unit Cost (₱)",
    "unit cost (php)": "Unit Cost (₱)",
    "unit cost (₱)": "Unit Cost (₱)",
    "unit price": "Unit Cost (₱)",
    "total cost": "Total Cost (₱)",
    "total cost (php)": "Total Cost (₱)",
    "total cost (₱)": "Total Cost (₱)",
    "total": "Total Cost (₱)",
    "category": "Category",
    "notes": "Notes",
    "store": "Store Name",
    "store name": "Store Name",
    "expense name": "Expense Name",
    "expense": "Expense Name",
    "amount": "Amount (₱)",
    "amount (php)": "Amount (₱)",
    "amount (₱)": "Amount (₱)",
}

HEADER_FONT = Font(bold=True)


# ---------------------------------------------------------------------------
# Blank frames
# ---------------------------------------------------------------------------

def blank_materials_df() -> pd.DataFrame:
    return pd.DataFrame(columns=CORE_MATERIAL_COLUMNS)


def blank_expenses_df() -> pd.DataFrame:
    return pd.DataFrame(columns=CORE_EXPENSE_COLUMNS)


# ---------------------------------------------------------------------------
# Validation / normalization
# ---------------------------------------------------------------------------

def _normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for col in df.columns:
        key = str(col).strip().lower()
        rename_map[col] = COLUMN_ALIASES.get(key, str(col).strip())
    return df.rename(columns=rename_map)


def validate_and_normalize_materials(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Normalize an uploaded/loaded materials sheet to the core schema.

    Raises ValueError if a truly required column is missing after alias
    resolution. Returns (normalized_df, warnings).
    """
    warnings: list[str] = []
    df = _normalize_headers(df.copy())
    df = df.drop(columns=[c for c in LEGACY_DROPPED_MATERIAL_COLUMNS if c in df.columns])

    missing_required = [c for c in REQUIRED_MATERIAL_COLUMNS if c not in df.columns]
    if missing_required:
        raise ValueError(
            "Missing required column(s): " + ", ".join(missing_required)
            + ". Download the template for the expected format."
        )

    for col in ("Category", "Notes"):
        if col not in df.columns:
            df[col] = ""
            warnings.append(f'Added missing optional column "{col}".')

    if "Store Name" not in df.columns:
        # Default to the "unknown provenance" sentinel, not a real store
        # name -- defaulting to e.g. "Plug and Go" would falsely assert
        # these rows came from that store and would wrongly make them
        # eligible for the Store -> Unit Cost auto-lookup in
        # apply_store_pricing() against a store they were never priced from.
        df["Store Name"] = catalog.CUSTOM_STORE_LABEL
        warnings.append('Added missing "Store Name" column (defaulted to "Custom / Manual").')

    if "Total Cost (₱)" not in df.columns:
        df["Total Cost (₱)"] = 0.0

    for col in NUMERIC_MATERIAL_COLUMNS:
        # astype(float) (not just fillna(0.0)) matters: a column that's all
        # whole numbers reads back from Excel as int64, which makes the
        # "accounting" display format silently drop decimal places.
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)

    df["Total Cost (₱)"] = df["Quantity"] * df["Unit Cost (₱)"]

    for col in ("Material", "Brand", "Unit of Measurement", "Category", "Notes", "Store Name"):
        df[col] = df[col].fillna("").astype(str)
    df.loc[df["Store Name"].str.strip() == "", "Store Name"] = catalog.CUSTOM_STORE_LABEL

    custom_cols = [c for c in df.columns if c not in CORE_MATERIAL_COLUMNS]
    df = df[CORE_MATERIAL_COLUMNS + custom_cols].reset_index(drop=True)
    return df, warnings


def validate_and_normalize_expenses(df: pd.DataFrame) -> pd.DataFrame:
    df = _normalize_headers(df.copy())
    for col in CORE_EXPENSE_COLUMNS:
        if col not in df.columns:
            df[col] = 0.0 if col in NUMERIC_EXPENSE_COLUMNS else ""
    for col in NUMERIC_EXPENSE_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)
    for col in ("Expense Name", "Notes"):
        df[col] = df[col].fillna("").astype(str)
    return df[CORE_EXPENSE_COLUMNS].reset_index(drop=True)


def recompute_material_totals(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ("Quantity", "Unit Cost (₱)"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)
    if "Quantity" in df.columns and "Unit Cost (₱)" in df.columns:
        df["Total Cost (₱)"] = df["Quantity"] * df["Unit Cost (₱)"]
    return df


def apply_store_pricing(materials_df: pd.DataFrame, catalog_df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """For every Materials row whose Store Name is a real, catalogued store
    (not the "Custom / Manual" sentinel, not blank), look up
    (Material, Store Name) against the catalog and overwrite Unit Cost (₱)
    with that store's price.

    Rows with no matching catalog entry are left with their existing
    Unit Cost (₱) untouched -- resetting to 0 would silently destroy a real,
    previously-entered price the moment a store name doesn't have a
    matching catalog row; a warning is returned instead so the caller can
    surface it without this module depending on Streamlit.

    Returns (updated_df, warnings).
    """
    df = materials_df.copy()
    warnings: list[str] = []
    if catalog_df is None or catalog_df.empty or "Store Name" not in df.columns:
        return df, warnings

    price_lookup = {
        (str(row["Material"]).strip(), str(row["Store Name"]).strip()): row["Unit Cost (₱)"]
        for _, row in catalog_df.iterrows()
    }

    for i, row in df.iterrows():
        store = str(row.get("Store Name", "")).strip()
        if not store or store == catalog.CUSTOM_STORE_LABEL:
            continue
        material = str(row.get("Material", "")).strip()
        key = (material, store)
        if key in price_lookup:
            df.at[i, "Unit Cost (₱)"] = float(price_lookup[key])
        else:
            warnings.append(f'No "{store}" catalog price found for "{material}" -- Unit Cost left as-is.')

    return df, warnings


def apply_material_categories(df: pd.DataFrame, catalog_df: pd.DataFrame) -> pd.DataFrame:
    """Category is derived from Material via the Category<->Material
    mapping the catalog itself already encodes (each catalog row pairs a
    Material with its Category) -- same "pick X, Y follows" pattern as
    apply_store_pricing() for Store -> Unit Cost, applied at the same
    point (right before recompute_material_totals, on every "Apply
    changes"). A Material with no catalog row (e.g. not yet added to the
    catalog, or a still-blank just-added row) keeps its current Category
    unchanged rather than getting blanked out."""
    df = df.copy()
    if catalog_df is None or catalog_df.empty or "Material" not in df.columns:
        return df

    category_lookup = catalog.material_category_map(catalog_df)

    for i, row in df.iterrows():
        material = str(row.get("Material", "")).strip()
        if material in category_lookup and category_lookup[material]:
            df.at[i, "Category"] = category_lookup[material]

    return df


def sort_materials_df(df: pd.DataFrame) -> pd.DataFrame:
    """Default browse order for the BOQ Materials table: Category, then
    Material, both ascending and case-insensitive (via a throwaway
    lowercase sort key, not a mutation of the real columns)."""
    if df is None or df.empty:
        return df
    df = df.copy()
    df["_cat"] = df["Category"].astype(str).str.lower()
    df["_mat"] = df["Material"].astype(str).str.lower()
    df = df.sort_values(["_cat", "_mat"], kind="mergesort")
    return df.drop(columns=["_cat", "_mat"])


def materials_total(df: pd.DataFrame) -> float:
    if df.empty or "Total Cost (₱)" not in df.columns:
        return 0.0
    return float(pd.to_numeric(df["Total Cost (₱)"], errors="coerce").fillna(0.0).sum())


def expenses_total(df: pd.DataFrame) -> float:
    if df.empty or "Amount (₱)" not in df.columns:
        return 0.0
    return float(pd.to_numeric(df["Amount (₱)"], errors="coerce").fillna(0.0).sum())


# ---------------------------------------------------------------------------
# Workbook building
# ---------------------------------------------------------------------------

def _write_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for _, row in df.iterrows():
        ws.append(list(row))
    ws.freeze_panes = "A2"
    for i, col in enumerate(df.columns, start=1):
        width = max(12, min(40, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_meta_sheet(ws, meta: dict) -> None:
    ws.append(["key", "value"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for key, value in meta.items():
        ws.append([key, value])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30


def build_workbook(materials_df: pd.DataFrame, expenses_df: pd.DataFrame, meta: dict) -> Workbook:
    wb = Workbook()
    ws_materials = wb.active
    ws_materials.title = "Materials"
    _write_sheet(ws_materials, materials_df)

    ws_expenses = wb.create_sheet("Expenses")
    _write_sheet(ws_expenses, expenses_df)

    ws_meta = wb.create_sheet("Meta")
    _write_meta_sheet(ws_meta, meta)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_catalog_dropdowns(
    wb: Workbook,
    category_list_df: pd.DataFrame,
    catalog_df: pd.DataFrame,
    max_row: int = 500,
) -> None:
    """Add Excel dropdown lists for Category and Material on the Materials
    sheet. Category comes from the editable Category master list; Material
    comes straight from the distinct Material names already in the catalog
    (there's no separate master list for it). Values are written to a
    hidden helper sheet and referenced by range (not as a literal list) so
    names with commas or quotes don't break anything, and there's no
    255-character formula length limit to worry about.
    """

    def _clean_values(series: pd.Series) -> list:
        # Drop actual nulls (NaN/None/pd.NA) *before* stringifying, not
        # after -- a list can transiently hold a still-blank row right
        # after the user clicks "+" to add one but hasn't filled it in yet,
        # and depending on dtype that blank cell isn't always a plain string
        # "nan" once stringified. dropna() reliably catches all of those.
        return sorted({str(v).strip() for v in series.dropna().tolist() if str(v).strip()})

    try:
        categories = _clean_values(category_list_df["Category"]) if category_list_df is not None and not category_list_df.empty else []
        materials_list = catalog.distinct_materials(catalog_df)
    except Exception:
        # Dropdown suggestions are a nice-to-have -- malformed/partial
        # list data should never block downloading the template itself.
        return
    if not categories and not materials_list:
        return

    lists_ws = wb.create_sheet("Lists")
    lists_ws["A1"] = "Category"
    lists_ws["B1"] = "Material"
    for i, value in enumerate(categories, start=2):
        lists_ws.cell(row=i, column=1, value=value)
    for i, value in enumerate(materials_list, start=2):
        lists_ws.cell(row=i, column=2, value=value)
    # Hidden, not deleted: openpyxl data validation formulas need the sheet
    # to still exist in the workbook, just out of the user's way.
    lists_ws.sheet_state = "hidden"

    materials_ws = wb["Materials"]
    cat_col = get_column_letter(CORE_MATERIAL_COLUMNS.index("Category") + 1)
    material_col = get_column_letter(CORE_MATERIAL_COLUMNS.index("Material") + 1)

    # showDropDown=False is not a typo -- in the underlying Excel file
    # format that flag means "suppress the dropdown arrow" when True, so it
    # must stay False for the arrow to actually show up.
    if categories:
        dv_category = DataValidation(
            type="list",
            formula1=f"=Lists!$A$2:$A${len(categories) + 1}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
        )
        dv_category.add(f"{cat_col}2:{cat_col}{max_row}")
        materials_ws.add_data_validation(dv_category)

    if materials_list:
        dv_material = DataValidation(
            type="list",
            formula1=f"=Lists!$B$2:$B${len(materials_list) + 1}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
        )
        dv_material.add(f"{material_col}2:{material_col}{max_row}")
        materials_ws.add_data_validation(dv_material)


def generate_template_workbook(
    category_list_df: pd.DataFrame | None = None,
    catalog_df: pd.DataFrame | None = None,
) -> bytes:
    """Build the downloadable blank BOQ template. Category/Material columns
    get Excel dropdown suggestions from the editable Category master list
    and the catalog's distinct Material names (typed values outside the
    list are still accepted -- these are suggestions, not a hard
    restriction, since a new project may need materials that aren't
    catalogued yet). Pass the current in-session data explicitly to
    reflect unsaved edits; defaults to what's on disk otherwise.
    """
    if category_list_df is None:
        category_list_df = catalog.load_category_list()
    if catalog_df is None:
        catalog_df = catalog.load_catalog()

    materials = pd.DataFrame(
        [
            {
                "Category": "Panels",
                "Material": "Solar Panel 550W Monocrystalline",
                "Brand": "JA Solar",
                "Quantity": 10,
                "Unit of Measurement": "pcs",
                "Store Name": "Plug and Go",
                "Unit Cost (₱)": 8500.00,
                "Total Cost (₱)": 85000.00,
                "Notes": "EXAMPLE ROW -- delete before use",
            }
        ],
        columns=CORE_MATERIAL_COLUMNS,
    )
    expenses = pd.DataFrame(
        [
            {
                "Expense Name": "Delivery / Logistics",
                "Amount (₱)": 5000.00,
                "Notes": "EXAMPLE ROW -- delete before use",
            }
        ],
        columns=CORE_EXPENSE_COLUMNS,
    )
    meta = {
        "project_name": "",
        "slug": "",
        "created_at": "",
        "last_modified_at": "",
        "schema_version": SCHEMA_VERSION,
    }
    wb = build_workbook(materials, expenses, meta)
    _add_catalog_dropdowns(wb, category_list_df, catalog_df)
    return workbook_to_bytes(wb)


def export_bytes(materials_df: pd.DataFrame, expenses_df: pd.DataFrame, meta: dict) -> bytes:
    materials_df = recompute_material_totals(materials_df)
    wb = build_workbook(materials_df, expenses_df, meta)
    return workbook_to_bytes(wb)


# ---------------------------------------------------------------------------
# Project persistence
# ---------------------------------------------------------------------------

def _slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-")
    return slug or "project"


def _unique_slug(base_slug: str) -> str:
    slug = base_slug
    n = 2
    while (PROJECTS_DIR / f"{slug}.xlsx").exists():
        slug = f"{base_slug}-{n}"
        n += 1
    return slug


def _project_path(slug: str) -> Path:
    return PROJECTS_DIR / f"{slug}.xlsx"


def _read_meta(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Meta"]
    meta = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            meta[str(row[0])] = row[1]
    wb.close()
    return meta


def list_projects() -> list[dict]:
    PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
    projects = []
    for path in sorted(PROJECTS_DIR.glob("*.xlsx")):
        try:
            meta = _read_meta(path)
        except Exception:
            continue
        projects.append(
            {
                "slug": path.stem,
                "name": meta.get("project_name") or path.stem,
                "path": path,
                "last_modified_at": meta.get("last_modified_at", ""),
            }
        )
    projects.sort(key=lambda p: str(p["name"]).lower())
    return projects


def load_project(slug: str) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    path = _project_path(slug)
    materials_df = pd.read_excel(path, sheet_name="Materials", engine="openpyxl")
    expenses_df = pd.read_excel(path, sheet_name="Expenses", engine="openpyxl")
    meta = _read_meta(path)
    materials_df, _ = validate_and_normalize_materials(materials_df)
    expenses_df = validate_and_normalize_expenses(expenses_df)
    return materials_df, expenses_df, meta


def save_project(slug: str, materials_df: pd.DataFrame, expenses_df: pd.DataFrame, meta: dict) -> Path:
    PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
    materials_df = recompute_material_totals(materials_df)
    meta = dict(meta)
    meta["last_modified_at"] = datetime.now().isoformat(timespec="seconds")
    meta.setdefault("schema_version", SCHEMA_VERSION)
    wb = build_workbook(materials_df, expenses_df, meta)
    path = _project_path(slug)
    wb.save(path)
    return path


def create_project(
    name: str,
    materials_df: pd.DataFrame | None = None,
    expenses_df: pd.DataFrame | None = None,
) -> str:
    base_slug = _slugify(name)
    slug = _unique_slug(base_slug)
    now = datetime.now().isoformat(timespec="seconds")
    meta = {
        "project_name": name.strip(),
        "slug": slug,
        "created_at": now,
        "last_modified_at": now,
        "schema_version": SCHEMA_VERSION,
    }
    materials_df = (
        recompute_material_totals(materials_df) if materials_df is not None else blank_materials_df()
    )
    expenses_df = expenses_df if expenses_df is not None else blank_expenses_df()
    save_project(slug, materials_df, expenses_df, meta)
    return slug


def read_uploaded_boq(file) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Parse an uploaded .xlsx file into (materials_df, expenses_df, warnings).

    Accepts either a plain single-sheet BOQ (first sheet treated as
    materials) or a file that already follows this app's Materials/Expenses
    sheet convention.
    """
    xls = pd.ExcelFile(file, engine="openpyxl")
    sheet_names = xls.sheet_names
    materials_sheet = "Materials" if "Materials" in sheet_names else sheet_names[0]
    materials_df = pd.read_excel(xls, sheet_name=materials_sheet)
    materials_df, warnings = validate_and_normalize_materials(materials_df)

    if "Expenses" in sheet_names:
        expenses_df = pd.read_excel(xls, sheet_name="Expenses")
        expenses_df = validate_and_normalize_expenses(expenses_df)
    else:
        expenses_df = blank_expenses_df()

    return materials_df, expenses_df, warnings
